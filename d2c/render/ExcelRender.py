#!/usr/bin/python
# encoding: utf-8

import csv
import os
import os.path

from openpyxl import Workbook, load_workbook

from ..varVo import TemplateInfo, ClassVo,RowData,VarVo


def gettitles(ws, rowIdx=1, genDict=False):
    max_col = ws.max_column
    max_row = ws.max_row
    titles = {} if genDict else []
    i = 1
    for row in ws.iter_rows(min_row=rowIdx, max_col=max_col, max_row=rowIdx):
        for cell in row:
            if genDict:
                titles[cell.value] = i
                i += 1
            else:
                titles.append(cell.value or '')
    return titles


def readfile(filename, colNames, kmap):
    wb = load_workbook(os.path.abspath(filename), read_only=True, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    # titlemap = gettitles(ws, 2, True)
    # for colName in colNames:
    #     colidx = titlemap[colName]

    for row in ws.iter_rows(min_row=1, max_row=max_row):
        yield row
        # for cell in col:
        #     v = cell.value
        #     if v:
        #         kmap[v] = True


class ExcelRender:
    def __init__(self, cls:ClassVo, defaultTemplates:[TemplateInfo], d2c):
        """
        """
        self.cls = cls
        self.defaultTemplates:[TemplateInfo] = defaultTemplates
        self.d2c = d2c
        self.templates = cls.templates if len(cls.templates) > 0 else defaultTemplates

        self.csvName = os.path.join(d2c._config.dataDir, cls.csvName)

    def exists(self):
        return os.path.isfile(self.csvName)

    def render(self):
        rows = []
        reader = readfile(self.csvName, None, None)
        line_num = 1
        for rawRow in reader:
            # rawRow 从第5行开始读取
            if line_num < 5:
                line_num += 1
                continue
            row = RowData(self.cls)
            rows.append(row)
            originCount = len(self.cls.originVars)
            for i in range(originCount):
                var = self.cls.originVars[i].clone()
                var.value = str(rawRow[i].value) if rawRow[i].value else ''
                row.originVars.append(var)
                if not var.isDel:
                    row.vars.append(var)
                    # checklink
            row.indexVars = [row.vars[i] for i in self.cls.indexs]
            # print self.cls.indexs,  row.indexValues

        return rows
