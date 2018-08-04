import os

from openpyxl import Workbook, load_workbook, worksheet

from ..config import Config
from ..varVo import ClassVo, RowData, VarVo


def is_skip_field(flag: str, data_filter: str) -> bool:
    if not data_filter:
        return False
    return flag.upper() not in data_filter


class ExcelParser:
    @classmethod
    def parse(cls, excelPath: str, config: Config) -> ClassVo:
        wb = load_workbook(os.path.abspath(excelPath), data_only=True)
        ws = wb.active
        max_col = ws.max_column

        clsVo = ClassVo()
        clsVo.csvName = os.path.basename(excelPath)
        clsVo.name = clsVo.csvName.rsplit(
            sep='.', maxsplit=2)[0].split(sep='_')[1]

        # read all vars
        for i in range(1, max_col + 1):
            name_cell = ws.cell(row=2, column=i)
            var_name = name_cell.value
            var_type = ws.cell(row=3, column=i).value
            exprot = ws.cell(row=4, column=i).value
            varvo = VarVo()
            clsVo.originVars.append(varvo)
            varvo.isDel = not var_name or not var_type
            if not varvo.isDel:
                varvo.isDel = is_skip_field(exprot, config.data_filter)
            if varvo.isDel:
                continue

            clsVo.vars.append(varvo)
            varvo.name = var_name
            varvo.type = var_type
            comment = name_cell.comment and name_cell.comment.text

        clsVo.indexNames.append(clsVo.vars[0].name)
        clsVo.indexs.append(0)
        return clsVo
